import openpyxl
import os
import shutil


def get_used_file(usage_path):
        
    todo_folder_path = os.path.join(usage_path)
    todo_files = [f for f in os.listdir(todo_folder_path) if f.endswith(".xlsx")]
    if not todo_files:
        print("No Excel files found in the 'Zuzyty Material' folder.")
        return
    todo_file_name = todo_files[0]
    todo_file_path = os.path.join(todo_folder_path, todo_file_name)
    return todo_file_path

def find_missing_records(database_path, usage_path):

    database_wb = openpyxl.load_workbook(database_path)

    
    usage_wb = openpyxl.load_workbook(get_used_file(usage_path))

    database_sheet = database_wb['Lagerbestand M0129']
    usage_sheet = usage_wb['Materialliste']
    missing_records = []

    for usage_row in usage_sheet.iter_rows(min_row=6, values_only=True):
        asset_name, _ = usage_row[1], usage_row[4]

        asset_found = False  
        for database_row in database_sheet.iter_rows(min_row=2, max_row=database_sheet.max_row, values_only=True):
            asset_number = database_row[1]

            if asset_number == asset_name:
                asset_found = True
                break

        # Check if asset_name was not found in the database
        if not asset_found and asset_name is not None:
            missing_records.append(asset_name)

    print(f"These records are missing in Lager: {missing_records}" )

def update_quantities(database_path, usage_path, done_folder_path):
    database_wb = openpyxl.load_workbook(database_path)

    
    usage_wb = openpyxl.load_workbook(get_used_file(usage_path))
    #usage_wb = openpyxl.load_workbook(usage_path)

    database_sheet = database_wb['Lagerbestand M0129']
    usage_sheet = usage_wb['Materialliste']
    #print(database_sheet)
    count = 0
    for row in usage_sheet.iter_rows(min_row=6, values_only=True):
        asset_name, used_quantity = row[1], row[4]
        #print(asset_name, used_quantity)
        #print(row)
        
        for index, database_row in enumerate(database_sheet.iter_rows(min_row=2, max_row=database_sheet.max_row, values_only=True), start=2):
            asset_number, current_quantity = database_row[1], database_row[3]
            
            if asset_number == asset_name:
                
                if current_quantity is not None and used_quantity is not None:
                    #new_quantity = max(current_quantity - used_quantity, 0)
                    new_quantity = current_quantity - used_quantity
                    modified_cell = database_sheet.cell(row=index, column=4)
                    modified_cell.value = new_quantity
                    count += 1
                    print(f"Row processed: {index}, Asset Number: {asset_number}")
            
    print(f"This many records were modified: {count}")

    print(f"File saving proces started for: {database_path}")            
    database_wb.save(database_path)
    print(f"File was saved under this path: {database_path}")
    
    done_file_path = os.path.join(done_folder_path) #Ogarnac jezeli juz plik istnieje. Unikatowe umie tworzyc czy cos - do zrobienia
    shutil.move(get_used_file(usage_path), done_file_path)

    print(f"Material file was finished and moved to the folder: {done_file_path}")



if __name__ == "__main__":
    print("Script starting")
    #Replace these paths with the actual paths to your Excel files
    database_file_path = "Lagerbestand-M0129.xlsx"
    usage_file_path = "Zuzyty Material"
    done_folder_path = "Gotowe"
    find_missing_records(database_file_path, usage_file_path)
    update_quantities(database_file_path, usage_file_path, done_folder_path)
    
    input("Press Enter to exit...")

#Ogarnac jezeli juz plik istnieje. Unikatowe umie tworzyc czy cos - do zrobienia
#Dodawanie
    #Alert ze jest minus
    #alerty bledow
    
    