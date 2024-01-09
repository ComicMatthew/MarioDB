from my_functions import get_used_file, show_windows_alert
from config import database_file_path, usage_subtraction_file_path, done_folder_path
import openpyxl
import os
import shutil

from datetime import datetime

def find_missing_records(database_path, usage_path):

    database_wb = openpyxl.load_workbook(database_path)
    todo_file_path, todo_file_name = get_used_file(usage_path)
    usage_wb = openpyxl.load_workbook(todo_file_path)

    database_sheet = database_wb['Lagerbestand M0129']
    usage_sheet = usage_wb['Materialliste']
    missing_records = []

    for usage_row in usage_sheet.iter_rows(min_row=6, values_only=True):
        asset_name, _ = usage_row[1], usage_row[4]

        asset_found = False  
        for database_row in database_sheet.iter_rows(min_row=2, max_row=database_sheet.max_row, values_only=True):
            asset_number = database_row[1]

            if asset_number == asset_name:
                asset_found = True
                break

        # Check if asset_name was not found in the database
        if not asset_found and asset_name is not None:
            missing_records.append(asset_name)

    print(f"Tych elementow nie ma w pliku magazynowym: {missing_records} a sa dostepne w pliku: {todo_file_name}" )

def update_quantities(database_path, usage_path, done_folder_path):
    try:
        database_wb = openpyxl.load_workbook(database_path)
        todo_file_path, todo_file_name = get_used_file(usage_path)
        usage_wb = openpyxl.load_workbook(todo_file_path)
        #usage_wb = openpyxl.load_workbook(usage_path)

        database_sheet = database_wb['Lagerbestand M0129']
        usage_sheet = usage_wb['Materialliste']
        #print(database_sheet)
        count = 0
        negative_record = []
        for row in usage_sheet.iter_rows(min_row=6, values_only=True):
            asset_name, used_quantity = row[1], row[4]
            #print(asset_name, used_quantity)
            #print(row)
            
            for index, database_row in enumerate(database_sheet.iter_rows(min_row=2, max_row=database_sheet.max_row, values_only=True), start=2):
                asset_number, current_quantity = database_row[1], database_row[3]
                
                if asset_number == asset_name:
                    
                    if current_quantity is not None and used_quantity is not None:
                        #new_quantity = max(current_quantity - used_quantity, 0)
                        new_quantity = current_quantity - used_quantity
                        modified_cell = database_sheet.cell(row=index, column=4)
                        modified_cell.value = new_quantity
                        count += 1
                        print(f"Wiersz zmieniony: {index}, Numer czesci: {asset_number}")
                        if new_quantity < 0:
                            print(f"!!!!! Elemnt z wiersza: {index} ma UJEMNA wartosc: {new_quantity}. Numer czesci: {asset_number} !!!")
                            negative_record.append(asset_number)
                            
        print(f"Liczba elementow ktore zostaly zedytowane: {count}")
        print(f"Te elementy na magazynie maja ujemna wartosc: {negative_record}")
        try:
            print(f"Porces zapisu pliku rozpoczety: {database_path}")            
            database_wb.save(database_path)
            print(f"Proces zapisu zakonczony: {database_path}")
            new_name = todo_file_name.split(".")[0] + datetime.now().strftime("%Y-%m-%d %H-%M-%S") + "." + todo_file_name.split(".")[1]
            done_file_path = os.path.join(done_folder_path, new_name)
            shutil.move(todo_file_path, done_file_path)
            #os.rename(todo_file_path, done_file_path)
            print(f"Plik z materialami obrobiony i przeniesiony do folderu: {done_file_path}")
        except PermissionError as e:
            print(f"Error: {e}")
            show_windows_alert("Ograniczony dostep do pliku", f"Baza danych nie zostala zapisana, gdyz dostep byl ograniczony. Prawdopodobnie Excel z baza danych jest otwarty. {str(e)}. Plik NIE zostal zapisany. Zamknij go i odpal skrypt ponownie")
        
        
    except shutil.Error as e:
        print(f"Error: {e}")
        show_windows_alert("W Folderze istnieje juz plik o tej nazwie", f"Baza danych Zostala zaktualizowana, wiec musisz tylko przeniesc plik do folderu '{done_file_path}': {str(e)}")



if __name__ == "__main__":
    print("Script starting")
    #Replace these paths with the actual paths to your Excel files
    find_missing_records(database_file_path, usage_subtraction_file_path)
    update_quantities(database_file_path, usage_subtraction_file_path, done_folder_path)
    
    input("Press Enter to exit...")

#Ogarnac jezeli juz plik istnieje. Unikatowe umie tworzyc czy cos - do zrobienia !! 
#Dodawanie
    #Alert ze jest minus X
    #alerty bledow X 
    
    