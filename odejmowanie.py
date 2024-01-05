import openpyxl
import os
import shutil
import ctypes

def show_windows_alert(title, message):
    ctypes.windll.user32.MessageBoxW(0, message, title, 1)


def get_used_file(usage_path):
        
    todo_folder_path = os.path.join(usage_path)
    todo_files = [f for f in os.listdir(todo_folder_path) if f.endswith(".xlsx")]
    if not todo_files:
        error_message = f"W Folderze {todo_folder_path} nie ma zadnego pliku do przetworzenia!"
        show_windows_alert("Brak Pliku", error_message)
        print(error_message)       
    else:
        todo_file_name = todo_files[0]
        todo_file_path = os.path.join(todo_folder_path, todo_file_name)
        return todo_file_path

def find_missing_records(database_path, usage_path):

    database_wb = openpyxl.load_workbook(database_path)

    usage_wb = openpyxl.load_workbook(get_used_file(usage_path))

    database_sheet = database_wb['Lagerbestand M0129']
    usage_sheet = usage_wb['Materialliste']
    missing_records = []

    for usage_row in usage_sheet.iter_rows(min_row=6, values_only=True):
        asset_name, _ = usage_row[1], usage_row[4]

        asset_found = False  
        for database_row in database_sheet.iter_rows(min_row=2, max_row=database_sheet.max_row, values_only=True):
            asset_number = database_row[1]

            if asset_number == asset_name:
                asset_found = True
                break

        # Check if asset_name was not found in the database
        if not asset_found and asset_name is not None:
            missing_records.append(asset_name)

    print(f"Tych elementow nie ma w pliku magazynowym: {missing_records}" )

def update_quantities(database_path, usage_path, done_folder_path):
    try:
        database_wb = openpyxl.load_workbook(database_path)

        usage_wb = openpyxl.load_workbook(get_used_file(usage_path))
        #usage_wb = openpyxl.load_workbook(usage_path)

        database_sheet = database_wb['Lagerbestand M0129']
        usage_sheet = usage_wb['Materialliste']
        #print(database_sheet)
        count = 0
        negative_record = []
        for row in usage_sheet.iter_rows(min_row=6, values_only=True):
            asset_name, used_quantity = row[1], row[4]
            #print(asset_name, used_quantity)
            #print(row)
            
            for index, database_row in enumerate(database_sheet.iter_rows(min_row=2, max_row=database_sheet.max_row, values_only=True), start=2):
                asset_number, current_quantity = database_row[1], database_row[3]
                
                if asset_number == asset_name:
                    
                    if current_quantity is not None and used_quantity is not None:
                        #new_quantity = max(current_quantity - used_quantity, 0)
                        new_quantity = current_quantity - used_quantity
                        modified_cell = database_sheet.cell(row=index, column=4)
                        modified_cell.value = new_quantity
                        count += 1
                        print(f"Wiersz zmieniony: {index}, Numer czesci: {asset_number}")
                        if new_quantity < 0:
                            print(f"!!!!! Elemnt z wiersza: {index} ma UJEMNA wartosc: {new_quantity}. Numer czesci: {asset_number} !!!")
                            negative_record.append(asset_number)
                            
        print(f"Liczba elementow ktore zostaly zedytowane: {count}")
        print(f"Te elementy na magazynie maja ujemna wartosc: {negative_record}")
        try:
            print(f"Porces zapisu pliku rozpoczety: {database_path}")            
            database_wb.save(database_path)
            print(f"Proces zapisu zakonczony: {database_path}")
            done_file_path = os.path.join(done_folder_path) #Ogarnac jezeli juz plik istnieje. Unikatowe umie tworzyc czy cos - do zrobienia
        
            shutil.move(get_used_file(usage_path), done_file_path)
            print(f"Plik z materialami obrobiony i przeniesiony do folderu: {done_file_path}")
        except PermissionError as e:
            print(f"Error: {e}")
            show_windows_alert("Ograniczony dostep do pliku", f"Baza danych nie zostala zapisana, gdyz dostep byl ograniczony. Prawdopodobnie Excel z baza danych jest otwarty. {str(e)}. Plik NIE zostal zapisany. Zamknij go i odpal skrypt ponownie")
        
        
    except shutil.Error as e:
        print(f"Error: {e}")
        show_windows_alert("W Folderze istnieje juz plik o tej nazwie", f"Baza danych Zostala zaktualizowana, wiec musisz tylko przeniesc plik do folderu '{done_file_path}': {str(e)}")



if __name__ == "__main__":
    print("Script starting")
    #Replace these paths with the actual paths to your Excel files
    database_file_path = "Lagerbestand-M0129.xlsx"
    usage_file_path = "Zuzyty Material"
    done_folder_path = "Gotowe"
    find_missing_records(database_file_path, usage_file_path)
    update_quantities(database_file_path, usage_file_path, done_folder_path)
    
    input("Press Enter to exit...")

#Ogarnac jezeli juz plik istnieje. Unikatowe umie tworzyc czy cos - do zrobienia
#Dodawanie
    #Alert ze jest minus X
    #alerty bledow X 
    
    